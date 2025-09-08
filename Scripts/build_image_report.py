#!/usr/bin/env python3
import os, argparse, math
import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ---------- report style ----------
mpl.rcParams.update({
    "figure.dpi": 180, "savefig.dpi": 300,
    "font.size": 11, "axes.titlesize": 12, "axes.labelsize": 11,
    "axes.linewidth": 0.8, "xtick.labelsize": 10, "ytick.labelsize": 10,
    "figure.facecolor": "white", "axes.facecolor": "white",
})

# ---------- helpers ----------
def to_NYX(a):
    a = np.squeeze(a)
    if a.ndim == 2: a = a[None, ...]
    elif a.ndim == 4 and a.shape[-1] == 1: a = a[..., 0]
    if a.ndim != 3: raise ValueError(f"Expected 2D/3D (or 4D with channel=1). Got {a.shape}")
    if a.shape[0] < max(a.shape[1], a.shape[2]): a = np.moveaxis(a, -1, 0)
    return a  # (N,Y,X)

def is_binaryish(img):
    u = np.unique(img)
    if u.size <= 4 and np.all(np.isin(u, [0,1,0.0,1.0])): return True
    return np.all((img >= -1e-6) & (img <= 1+1e-6))

def porosity(mask):
    m = (mask > 0.5) if not is_binaryish(mask) else (mask > 0.5)
    return float(m.mean()), int(m.sum()), int(m.size - m.sum())

def stats1d(a):
    a = np.asarray(a).ravel(); a = a[np.isfinite(a)]
    return dict(
        min=float(np.min(a)), max=float(np.max(a)),
        mean=float(np.mean(a)), std=float(np.std(a)),
        p5=float(np.percentile(a,5)), p50=float(np.percentile(a,50)), p95=float(np.percentile(a,95))
    )

def per_image_stats(vx_i, vy_i, den_i):
    sp = np.hypot(vx_i, vy_i)
    row = {}
    row.update({f"vx_{k}": v for k,v in stats1d(vx_i).items()})
    row.update({f"vy_{k}": v for k,v in stats1d(vy_i).items()})
    row.update({f"speed_{k}": v for k,v in stats1d(sp).items()})
    if den_i is not None:
        phi, void_px, solid_px = porosity(den_i)
        row["porosity"] = phi
        row["void_pixels"] = void_px
        row["solid_pixels"] = solid_px
    return row

def net_flow_direction_deg(vx_i, vy_i):
    vx_mean, vy_mean = float(np.mean(vx_i)), float(np.mean(vy_i))
    return float(np.degrees(np.arctan2(vy_mean, vx_mean)))

def render_pair(mask, vx, vy, display_idx_1based, out_png, global_vmax=None):
    """Two panels, like the paper (Input / Predicted |v|)."""
    H, W = vx.shape
    sp = np.hypot(vx, vy)
    vmax = float(global_vmax if global_vmax is not None else np.nanpercentile(sp, 99))
    vmax = vmax if vmax > 0 else 1.0

    fig, axs = plt.subplots(1, 2, figsize=(8.6, 4.6), constrained_layout=True)

    # Left: Input
    if mask is not None:
        im0 = axs[0].imshow(mask, origin="lower", cmap="gray", vmin=0, vmax=1, interpolation="nearest")
        cb0 = fig.colorbar(im0, ax=axs[0], fraction=0.046, pad=0.04); cb0.set_label("Input (0–1)")
    else:
        axs[0].text(0.5, 0.5, "den.npy not found", ha="center", va="center", transform=axs[0].transAxes)
    axs[0].set_title("Input")
    ticks = [0,32,64,96,128] if W==128 else [0, W//4, W//2, 3*W//4, W]
    axs[0].set_xlim(0, W); axs[0].set_ylim(0, H)
    axs[0].set_xticks(ticks); axs[0].set_yticks(ticks)

    # Right: Predicted |v|
    im1 = axs[1].imshow(sp, origin="lower", cmap="turbo", vmin=0, vmax=vmax, interpolation="nearest")
    import matplotlib.ticker as mticker
    cb1 = fig.colorbar(im1, ax=axs[1], fraction=0.046, pad=0.04)
    cb1.formatter = mticker.ScalarFormatter(useMathText=True); cb1.formatter.set_powerlimits((-3, 3)); cb1.update_ticks()
    cb1.set_label("|v|")
    axs[1].set_title("Predicted")
    axs[1].set_xlim(0, W); axs[1].set_ylim(0, H)
    axs[1].set_xticks(ticks); axs[1].set_yticks(ticks)

    head = f"Sample {display_idx_1based}"
    if mask is not None:
        phi, _, _ = porosity(mask)
        head += f"   •   φ={phi:.3f}"
    fig.suptitle(head, y=0.995, fontsize=12)

    os.makedirs(os.path.dirname(out_png), exist_ok=True)
    fig.savefig(out_png, bbox_inches="tight")
    plt.close(fig)

def write_sample_sheet(writer, sheet_name, vx, vy, den=None, vel_scale=None):
    H, W = vx.shape
    sp = np.hypot(vx, vy)

    # Pick a nice scale so values are ~0.1–10
    if vel_scale is None:
        p95 = float(np.percentile(sp, 95))
        if p95 > 0:
            exp = int(np.round(-np.log10(p95)))  # e.g., p95≈1e-4 -> exp=4
            vel_scale = 10.0 ** exp
        else:
            vel_scale = 1.0

    # Summary (original units)
    phi = void = solid = None
    if den is not None:
        phi, void, solid = porosity(den)
    summary = pd.DataFrame([
        ["Porosity φ", phi], ["Void pixels", void], ["Solid pixels", solid],
        ["Mean v_x", float(np.mean(vx))], ["Mean v_y", float(np.mean(vy))],
        ["Net flow angle (deg)", net_flow_direction_deg(vx, vy)],
        ["Mean |v|", float(np.mean(sp))], ["p95 |v|", float(np.percentile(sp,95))],
        ["Max |v|", float(np.max(sp))],
        ["Note", f"Matrices below are scaled by ×{vel_scale:g}"]
    ], columns=["Metric","Value"])
    summary.to_excel(writer, sheet_name=sheet_name, startrow=0, startcol=0, index=False)

    # Scaled matrices
    r = 14
    pd.DataFrame([[f"v_x (Y×X)  ×{vel_scale:g}"]]).to_excel(writer, sheet_name=sheet_name,
                                                           startrow=r, startcol=0, index=False, header=False)
    pd.DataFrame(vx * vel_scale).to_excel(writer, sheet_name=sheet_name, startrow=r+1, startcol=0,
                                          index=False, header=False)

    r = r + 1 + H + 2
    pd.DataFrame([[f"v_y (Y×X)  ×{vel_scale:g}"]]).to_excel(writer, sheet_name=sheet_name,
                                                           startrow=r, startcol=0, index=False, header=False)
    pd.DataFrame(vy * vel_scale).to_excel(writer, sheet_name=sheet_name, startrow=r+1, startcol=0,
                                          index=False, header=False)

    r = r + 1 + H + 2
    pd.DataFrame([[f"|v| (Y×X)  ×{vel_scale:g}"]]).to_excel(writer, sheet_name=sheet_name,
                                                            startrow=r, startcol=0, index=False, header=False)
    pd.DataFrame(sp * vel_scale).to_excel(writer, sheet_name=sheet_name, startrow=r+1, startcol=0,
                                          index=False, header=False)

def embed_figures_in_sheets(xlsx_path, mapping):
    """mapping: {sheet_name: figure_path}. Embeds each figure near the top-right."""
    wb = load_workbook(xlsx_path)
    for sheet, img_path in mapping.items():
        if not os.path.exists(img_path): continue
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        img = XLImage(img_path)
        ws.add_image(img, "J2")
    wb.save(xlsx_path)

# ---------- main ----------
def main():
    ap = argparse.ArgumentParser(description="Report-style Input/Predicted figures + per-sample velocity sheets (1–10).")
    ap.add_argument("--data-dir", default="Datasets")
    ap.add_argument("--out-dir", default="Results/flow_report",
                    help="Output folder that will contain the Excel and figs/")
    # If provided and relative, it's placed under --out-dir. If omitted, defaults to <out-dir>/flow_summary.xlsx
    ap.add_argument("--out-xlsx", default=None)
    args = ap.parse_args()

    # Resolve output locations (robust)
    out_dir_final = os.path.normpath(args.out_dir) if args.out_dir else "Results/flow_report"
    if args.out_xlsx:
        if os.path.isabs(args.out_xlsx):
            out_xlsx_path = os.path.normpath(args.out_xlsx)
            out_dir_final = os.path.dirname(out_xlsx_path)
        else:
            out_xlsx_path = os.path.join(out_dir_final, args.out_xlsx)
    else:
        out_xlsx_path = os.path.join(out_dir_final, "flow_summary.xlsx")

    os.makedirs(out_dir_final, exist_ok=True)
    figs_dir = os.path.join(out_dir_final, "figs")
    os.makedirs(figs_dir, exist_ok=True)

    # Load arrays
    den = None
    den_path = os.path.join(args.data_dir, "den.npy")
    if os.path.exists(den_path): den = to_NYX(np.load(den_path, allow_pickle=False))
    vx = to_NYX(np.load(os.path.join(args.data_dir, "vx.npy"), allow_pickle=False))
    vy = to_NYX(np.load(os.path.join(args.data_dir, "vy.npy"), allow_pickle=False))
    N, H, W = vx.shape
    if vy.shape != (N,H,W): raise SystemExit("vx/vy shape mismatch")

    # Global stats & color scale
    picks = list(range(min(10, N)))                  # samples 1..10
    global_vmax = float(np.nanpercentile(np.hypot(vx,vy), 99))

    # Per-image stats for all samples
    per_rows = []
    for i in range(N):
        den_i = den[i] if den is not None else None
        row = {"sample": i+1, **per_image_stats(vx[i], vy[i], den_i)}
        per_rows.append(row)
    per_df = pd.DataFrame(per_rows).sort_values("sample").reset_index(drop=True)

    # Render figures for samples 1–10
    figure_paths = {}
    for i in picks:
        fig_path = os.path.join(figs_dir, f"sample_{i+1:02d}.png")
        render_pair(den[i] if den is not None else None, vx[i], vy[i], i+1, fig_path, global_vmax=global_vmax)
        figure_paths[f"S{i+1:02d}"] = fig_path

    # Summary sheet
    summary_rows = [
        ["Total samples (N)", N],
        ["Grid (H×W)", f"{H}×{W}"],
        ["Figures/pages generated", f"{len(picks)} (Samples 1–10)"],
        ["Notes", "Each sample page (S01..S10) embeds its figure and shows summary + full matrices v_x, v_y, |v|."]
    ]
    if "porosity" in per_df:
        summary_rows.append(["Mean φ (all samples)", float(per_df["porosity"].mean())])
    summary_rows += [
        ["Mean |v| (all samples)", float(per_df["speed_mean"].mean())],
        ["Across-sample |v| 5–95%", f"{float(per_df['speed_p5'].mean()):.3g} – {float(per_df['speed_p95'].mean()):.3g}"]
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["Item","Value"])

    # Write Excel + one sheet per sample, numbered S01..S10
    with pd.ExcelWriter(out_xlsx_path, engine="openpyxl") as xl:
        summary_df.to_excel(xl, index=False, sheet_name="Summary")
        per_df.to_excel(xl, index=False, sheet_name="PerImageStats")
        for i in picks:
            sheet = f"S{i+1:02d}"
            den_i = den[i] if den is not None else None
            write_sample_sheet(xl, sheet, vx[i], vy[i], den_i)

    # Embed the figures into their corresponding sample sheets
    embed_figures_in_sheets(out_xlsx_path, figure_paths)

    print(f"Excel: {out_xlsx_path}")
    print("Sheets: Summary, PerImageStats, and S01..S{:02d}".format(len(picks)))
    print("Figures saved under:", figs_dir)

if __name__ == "__main__":
    main()
