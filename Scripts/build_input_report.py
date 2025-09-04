#!/usr/bin/env python3
"""
Build an input-only report from an existing dataset, mirroring the look/flow
of build_image_report but **without** velocity/prediction panels.

Enhancements:
- Sample title and porosity are on separate lines (no overlap).
- Outputs everything under a dedicated subfolder inside Results
  (default: Results/inputs_report/).
"""

import os
import argparse
import numpy as np
import pandas as pd
import matplotlib as mpl
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ---------- styling ----------
mpl.rcParams.update({
    "figure.dpi": 180, "savefig.dpi": 300,
    "font.size": 11, "axes.titlesize": 12, "axes.labelsize": 11,
    "axes.linewidth": 0.8, "xtick.labelsize": 10, "ytick.labelsize": 10,
    "figure.facecolor": "white", "axes.facecolor": "white",
})

# ---------- helpers ----------
def to_NYX(a: np.ndarray) -> np.ndarray:
    """Return array as (N, Y, X). Accepts (H,W), (N,H,W), or (H,W,N)."""
    a = np.asarray(a)
    a = np.squeeze(a)
    if a.ndim == 2:               # (H, W)
        a = a[None, ...]          # -> (1, H, W)
    elif a.ndim == 4 and a.shape[-1] == 1:
        a = a[..., 0]             # drop trailing channel
    if a.ndim != 3:
        raise ValueError(f"den.npy should be 2D/3D (or 4D with channel=1). Got {a.shape}")

    # If it's (H, W, N), move last axis to front.
    # Heuristic: if the last axis is much smaller/larger than the first two equal-ish dims (square grid),
    # treat last axis as N.
    y, x, maybe_n = a.shape[0], a.shape[1], a.shape[2]
    if y == x and maybe_n != y:
        a = np.moveaxis(a, -1, 0)  # (N, Y, X)
    # Otherwise assume it's already (N, Y, X)
    return a


def is_binaryish(img: np.ndarray) -> bool:
    u = np.unique(img)
    if u.size <= 4 and np.all(np.isin(u, [0, 1, 0.0, 1.0])):
        return True
    return np.all((img >= -1e-6) & (img <= 1 + 1e-6))


def porosity(mask: np.ndarray) -> float:
    m = (mask > 0.5) if not is_binaryish(mask) else (mask > 0.5)
    return float(m.mean())


def render_input(mask: np.ndarray, display_idx_1based: int, out_png: str):
    H, W = mask.shape
    # Manual layout so titles never overlap
    fig, ax = plt.subplots(1, 1, figsize=(4.8, 4.8), constrained_layout=False)
    ax.imshow(mask, origin="lower", cmap="gray", vmin=0, vmax=1, interpolation="nearest")

    # Axes cosmetics
    ticks = [0, 32, 64, 96, 128] if W == 128 else [0, W//4, W//2, 3*W//4, W]
    ax.set_xlim(0, W); ax.set_ylim(0, H)
    ax.set_xticks(ticks); ax.set_yticks(ticks)

    # Titles: put sample in suptitle, porosity in axes title with padding
    phi = porosity(mask)
    fig.suptitle(f"Sample {display_idx_1based}", y=0.98, fontsize=12)
    ax.set_title(f"Input (mask) • φ={phi:.3f}", pad=10)

    # Reserve space for suptitle
    fig.tight_layout(rect=[0, 0, 1, 0.95])

    os.makedirs(os.path.dirname(out_png), exist_ok=True)
    fig.savefig(out_png, bbox_inches="tight")
    plt.close(fig)


def embed_figures_in_sheets(xlsx_path: str, mapping: dict):
    wb = load_workbook(xlsx_path)
    for sheet, img_path in mapping.items():
        if not os.path.exists(img_path):
            continue
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        img = XLImage(img_path)
        ws.add_image(img, "J2")
    wb.save(xlsx_path)

# ---------- main ----------
def main():
    ap = argparse.ArgumentParser(description="Input-only report (den.npy) with embedded figures for S01..S10.")
    ap.add_argument("--data-dir", default="Datasets")
    ap.add_argument("--out-dir", default="Results/inputs_report",
                    help="Output folder that will contain the Excel and figs/")
    # Optional: if provided as a relative path, it's placed under --out-dir.
    # If omitted, defaults to <out-dir>/input_only_summary.xlsx
    ap.add_argument("--out-xlsx", default=None)
    ap.add_argument("--max-pages", type=int, default=10, help="How many sample sheets/figures to render (1..N)")
    args = ap.parse_args()

    # ---- Resolve output locations (robust; no double-join) ----
    out_dir_final = os.path.normpath(args.out_dir) if args.out_dir else "Results/inputs_report"
    if args.out_xlsx:
        if os.path.isabs(args.out_xlsx):
            out_xlsx_path = os.path.normpath(args.out_xlsx)
            out_dir_final = os.path.dirname(out_xlsx_path)
        else:
            out_xlsx_path = os.path.join(out_dir_final, args.out_xlsx)
    else:
        out_xlsx_path = os.path.join(out_dir_final, "input_only_summary.xlsx")

    os.makedirs(out_dir_final, exist_ok=True)
    figs_dir = os.path.join(out_dir_final, "figs")
    os.makedirs(figs_dir, exist_ok=True)

    # ---- Load den.npy (required) ----
    den_path = os.path.join(args.data_dir, "den.npy")
    if not os.path.exists(den_path):
        raise SystemExit(f"Missing {den_path}. This script only renders inputs from den.npy.")

    den = to_NYX(np.load(den_path, allow_pickle=False))
    N, H, W = den.shape

    # ---- Per-image porosity stats for all samples ----
    rows = [{"sample": i + 1, "porosity": porosity(den[i]), "height": H, "width": W} for i in range(N)]
    per_df = pd.DataFrame(rows).sort_values("sample").reset_index(drop=True)

    # ---- Render figures for S01..S{max_pages} ----
    pages = int(min(max(args.max_pages, 1), N))
    figure_paths = {}
    for i in range(pages):
        fig_path = os.path.join(figs_dir, f"sample_{i+1:02d}.png")
        render_input(den[i], i + 1, fig_path)
        figure_paths[f"S{i+1:02d}"] = fig_path

    # ---- Summary sheet ----
    summary_rows = [
        ["Total samples (N)", N],
        ["Grid (H×W)", f"{H}×{W}"],
        ["Figures/pages generated", f"{pages} (Samples 1–{pages:02d})"],
        ["Notes", "Each sample page (S01..SXX) embeds its input-only figure and porosity."],
        ["Mean φ (all samples)", float(per_df["porosity"].mean())],
    ]
    summary_df = pd.DataFrame(summary_rows, columns=["Item", "Value"])

    # ---- Write Excel ----
    with pd.ExcelWriter(out_xlsx_path, engine="openpyxl") as xl:
        summary_df.to_excel(xl, index=False, sheet_name="Summary")
        per_df.to_excel(xl, index=False, sheet_name="PerImageStats")
        for i in range(pages):
            sheet = f"S{i+1:02d}"
            sd = pd.DataFrame([["Porosity φ", float(per_df.loc[i, "porosity"])]], columns=["Metric", "Value"])
            sd.to_excel(xl, index=False, sheet_name=sheet, startrow=0, startcol=0)

    # ---- Embed figures ----
    embed_figures_in_sheets(out_xlsx_path, figure_paths)

    print(f"Excel: {out_xlsx_path}")
    print("Sheets: Summary, PerImageStats, and S01..S{:02d}".format(pages))
    print("Figures saved under:", figs_dir)


if __name__ == "__main__":
    main()
